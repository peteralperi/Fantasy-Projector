### LETS FUCKING GOOOOOOOOOOOOOOOOOOOOO

import openpyxl


wb = openpyxl.load_workbook("bigboy.xlsx") # brings excel file into wb

QBws = wb['QB'] # importing sheets from wb into 'ws
RBws = wb['RB']
WRws = wb['WR']
TEws = wb['TE']
Kws = wb['K']
DEFws = wb['Defense']
SCHEDws = wb['Schedule']


player = input("Enter Player Name: ") #getting player from user
team = input("Enter Player Team Abreviation: ")
pos = input("Enter Player Position: ")
week = input("Enter Current Week: ")
week = int(week)

#lists to use in math
pyds = []
pTds = []
ints = []
ryds = []
rTds = []
fumbs = []
games = []
recs = []
recyds = []
recTds = []
fgs = []
fga = []
onenine = []
twonine = []
threenine = []
fournine = []
fiveplus = []
xpt = []
def_pass_yds = []
def_ints = []
def_ptds = []
def_ryds = []
def_rtds = []
add_on = 0



if (pos == "QB") or (pos == "Qb") or (pos == "qb"):
    for column in QBws.iter_cols():  
        for cell in column:
            if cell.value == player:
                    for i in range(1,2):
                        
                        
                        pyds.append(QBws.cell(row=cell.row, column=i+1).value)
                        pTds.append(QBws.cell(row=cell.row, column=i+2).value)
                        ints.append(QBws.cell(row=cell.row, column=i+3).value)
                        ryds.append(QBws.cell(row=cell.row, column=i+4).value)
                        rTds.append(QBws.cell(row=cell.row, column=i+5).value)
                        fumbs.append(QBws.cell(row=cell.row, column=i+6).value)
                        games.append(QBws.cell(row=cell.row, column=i+7).value)
    sumgames = sum(games)
    base_proj = ((((sum(pyds))/sumgames)/25) + (((sum(pTds))/sumgames)*4) + (((sum(ints))/sumgames)*(-2)) + (((sum(ryds))/sumgames)/10) - (((sum(fumbs))/sumgames)*(2)))
    print(round(base_proj, 2))
    
    
    

if (pos == "RB") or (pos == "Rb") or (pos == "rb"):
    for column in RBws.iter_cols():  
        for cell in column:
            if cell.value == player:
                    for i in range(1,2):
                        
                        
                        ryds.append(RBws.cell(row=cell.row, column=i+1).value)
                        rTds.append(RBws.cell(row=cell.row, column=i+2).value)
                        recs.append(RBws.cell(row=cell.row, column=i+3).value)
                        recyds.append(RBws.cell(row=cell.row, column=i+4).value)
                        recTds.append(RBws.cell(row=cell.row, column=i+5).value)
                        fumbs.append(RBws.cell(row=cell.row, column=i+6).value)
                        games.append(RBws.cell(row=cell.row, column=i+7).value)
    sumgames = sum(games)
    base_proj =( (((sum(ryds))/sumgames)/10) + (((sum(rTds))/sumgames)*6) + (((sum(recs))/sumgames)) + (((sum(recyds))/sumgames)/10) + (((sum(recTds))/sumgames)*6) + (((sum(fumbs))/sumgames)*(-2)) )
    print(round(base_proj, 2))
    
    
             
if (pos == "WR") or (pos == "Wr") or (pos == "wr"):
    for column in WRws.iter_cols():  
        for cell in column:
            if cell.value == player:
                    for i in range(1,2):
                        
                        
                        recs.append(WRws.cell(row=cell.row, column=i+1).value)
                        recyds.append(WRws.cell(row=cell.row, column=i+2).value)
                        recTds.append(WRws.cell(row=cell.row, column=i+3).value)
                        ryds.append(WRws.cell(row=cell.row, column=i+4).value)
                        rTds.append(WRws.cell(row=cell.row, column=i+5).value)
                        fumbs.append(WRws.cell(row=cell.row, column=i+6).value)
                        games.append(WRws.cell(row=cell.row, column=i+7).value)
    sumgames = sum(games)
    base_proj = (((sum(recs))/sumgames)) + (((sum(recyds))/sumgames)/10) + (((sum(recTds))/sumgames)*6) + (((sum(ryds))/sumgames)/10) + (((sum(rTds))/sumgames)*6) + (((sum(fumbs))/sumgames)*(-2))
    print(round(base_proj, 2))

if (pos == "TE") or (pos == "Te") or (pos == "te"):
    for column in TEws.iter_cols():  
        for cell in column:
            if cell.value == player:
                    for i in range(1,2):
                        
                        
                        recs.append(TEws.cell(row=cell.row, column=i+1).value)
                        recyds.append(TEws.cell(row=cell.row, column=i+2).value)
                        recTds.append(TEws.cell(row=cell.row, column=i+3).value)
                        ryds.append(TEws.cell(row=cell.row, column=i+4).value)
                        rTds.append(TEws.cell(row=cell.row, column=i+5).value)
                        fumbs.append(TEws.cell(row=cell.row, column=i+6).value)
                        games.append(TEws.cell(row=cell.row, column=i+7).value)
    sumgames = sum(games)
    base_proj = (((sum(recs))/sumgames)) + (((sum(recyds))/sumgames)/10) + (((sum(recTds))/sumgames)*6) + (((sum(ryds))/sumgames)/10) + (((sum(rTds))/sumgames)*6) + (((sum(fumbs))/sumgames)*(-2))                            
    print(round(base_proj, 2))

if (pos == "K") or (pos == "k"):
    for column in Kws.iter_cols():  
        for cell in column:
            if cell.value == player:
                    for i in range(1,2):
                        
                        
                        fgs.append(Kws.cell(row=cell.row, column=i+1).value)
                        fga.append(Kws.cell(row=cell.row, column=i+2).value)
                        onenine.append(Kws.cell(row=cell.row, column=i+3).value)
                        twonine.append(Kws.cell(row=cell.row, column=i+4).value)
                        threenine.append(Kws.cell(row=cell.row, column=i+5).value)
                        fournine.append(Kws.cell(row=cell.row, column=i+6).value)
                        fiveplus.append(Kws.cell(row=cell.row, column=i+7).value)
                        xpt.append(Kws.cell(row=cell.row, column=i+8).value)
                        games.append(Kws.cell(row=cell.row, column=i+9).value)
    sumgames = sum(games)
    fg_missed = (sum(fga) - sum(fgs))
    base_proj = ((fg_missed/sumgames)*-2) + ((sum(onenine)/sumgames)*3) + ((sum(twonine)/sumgames)*3) + ((sum(threenine)/sumgames)*3) + ((sum(fournine)/sumgames)*4) + ((sum(fiveplus)/sumgames)*5) + (sum(xpt)/sumgames)
    print(round(base_proj, 2))
    
for row in SCHEDws.iter_rows(max_col=1):
    for cell in row:
        if cell.value == team:
            opponent = SCHEDws.cell(row = cell.row, column = week+1).value
str(opponent)            
opponent = opponent.replace('@', '')
          

#converting abbrevations to team name
if opponent == 'ARI':
    opponent = 'Cardinals'
if opponent == 'ATL':
    opponent = 'Falcons'
if opponent == 'BAL':
    opponent = 'Ravens'
if opponent == 'BUF':
    opponent = 'Bills'
if opponent == 'CAR':
    opponent = 'Panthers'
if opponent == 'CHI':
    opponent = 'Bears'
if opponent == 'CIN':
    opponent = 'Bengals'   
if opponent == 'CLE':
    opponent = 'Browns'
if opponent == 'DAL':
    opponent = 'Cowboys'   
if opponent == 'DEN':
    opponent = 'Broncos'
if opponent == 'DET':
    opponent = 'Lions'  
if opponent == 'GB':
    opponent = 'Packers'
if opponent == 'HOU':
    opponent = 'Texans'
if opponent == 'IND':
    opponent = 'Colts'
if opponent == 'JAC':
    opponent = 'Jaguars'
if opponent == 'KC':
    opponent = 'Chiefs'
if opponent == 'LAC':
    opponent = 'Chargers'
if opponent == 'LAR':
    opponent = 'Rams'
if opponent == 'LV':
    opponent = 'Raiders'
if opponent == 'MIA':
    opponent = 'Dolphins'
if opponent == 'MIN':
    opponent = 'Vikings'
if opponent == 'NE':
    opponent = 'Patriots'
if opponent == 'NO':
    opponent = 'Saints'
if opponent == 'NYG':
    opponent = 'Giants'
if opponent == 'NYJ':
    opponent = 'Jets'
if opponent == 'PHI':
    opponent = 'Eagles'    
if opponent == 'PIT':
    opponent = 'Steelers'
if opponent == 'SEA':
    opponent = 'Seahawks'
if opponent == 'SF':
    opponent = '49ers'
if opponent == 'TB':
    opponent = 'Buccaneers'
if opponent == 'TEN':
    opponent = 'Titans'
if opponent == 'WC':
    opponent = 'Commanders'
    

if (pos == "QB") or (pos == "Qb") or (pos == "qb") or (pos == "WR") or (pos == "Wr") or (pos == "wr") or (pos == "TE") or (pos == "Te") or (pos == "te"):
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_pass_yds.append(DEFws.cell(row= cell.row, column=2).value)
            break
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_ptds.append(DEFws.cell(row= cell.row, column=3).value)
            break
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_ints.append(DEFws.cell(row= cell.row, column=4).value)
            break
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_ryds.append(DEFws.cell(row= cell.row, column=9).value)
            break
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_rtds.append(DEFws.cell(row= cell.row, column=11).value)
            break
    
    def_pass_yds = (sum(def_pass_yds)/32)
    def_ptds = (sum(def_ptds)/32)
    def_ints = (sum(def_ints)/32)
    def_rush_yds = (sum(def_ryds)/32)
    def_rush_tds = (sum(def_rtds)/32)
    
    for row in DEFws.iter_rows(max_col=1):
        for cell in row:
            if cell.value == opponent:
                opp_passyds = DEFws.cell(row = cell.row, column = 2).value
                opp_passtds = DEFws.cell(row = cell.row, column = 3).value
                opp_passints = DEFws.cell(row = cell.row, column = 4).value
                
                
    for row in DEFws.iter_rows(min_col= 7, max_col=7):
        for cell in row:
            if cell.value == opponent:            
                opp_rushyds = DEFws.cell(row = cell.row, column = 9).value
                opp_rushtds = DEFws.cell(row = cell.row, column = 11).value
                
    
    diff_of_pyds = opp_passyds - def_pass_yds
    if (diff_of_pyds >=-75) and (diff_of_pyds <=75): 
        add_on = (diff_of_pyds / 25)
    if (diff_of_pyds > 75):
        add_on += 3
    if (diff_of_pyds < -75):
        add_on -= 3
    
    
    diff_of_ptds = opp_passtds - def_ptds # if the difference of passing touchdowns (compared to league average) is in a certain range, add a certain amt of points
    if (diff_of_ptds >= 1) and (diff_of_ptds<2):
        add_on +=.3
    if (diff_of_ptds <= -1) and (diff_of_ptds>-2):
        add_on -=.3
    if (diff_of_ptds > 2) and (diff_of_ptds<3):
        add_on +=.8
    if (diff_of_ptds <-2) and (diff_of_ptds>-3):
        add_on -=.8
    if (diff_of_ptds > 3):
        add_on += 1.4
    if (diff_of_ptds < -3):
        add_on -= 1.4
    
    if (pos == "QB") or (pos == "Qb") or (pos == "qb"):
        diff_of_ints = opp_passints - def_ints
        if (diff_of_ints >= 5):
            add_on -= 2
        if (diff_of_ints <= -5):
            add_on += 2
    if (pos == "QB") or (pos == "Qb") or (pos == "qb"):
        diff_of_rtds = opp_rushtds - def_rush_tds
        if (diff_of_rtds > 5):
            add_on += 1
        if (diff_of_rtds < -5):
            add_on -= 1
        
    final_proj = add_on + base_proj    




if (pos == "RB") or (pos == "Rb") or (pos == "rb"):
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_pass_yds.append(DEFws.cell(row= cell.row, column=2).value)
            break
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_ptds.append(DEFws.cell(row= cell.row, column=3).value)
            break
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_ryds.append(DEFws.cell(row= cell.row, column=9).value)
            break
    for column in DEFws.iter_rows(min_row=2):
        for cell in column:
            def_rtds.append(DEFws.cell(row= cell.row, column=11).value)
            break
    
    
    def_pass_yds = (sum(def_pass_yds)/32)
    def_ptds = (sum(def_ptds)/32)
    def_rush_yds = (sum(def_ryds)/32)
    def_rush_tds = (sum(def_rtds)/32)
    
    for row in DEFws.iter_rows(max_col=1):
        for cell in row:
            if cell.value == opponent:
                opp_passyds = DEFws.cell(row = cell.row, column = 2).value
                opp_passtds = DEFws.cell(row = cell.row, column = 3).value
                
                
    for row in DEFws.iter_rows(min_col= 7, max_col=7):
        for cell in row:
            if cell.value == opponent:            
                opp_rushyds = DEFws.cell(row = cell.row, column = 9).value
                opp_rushtds = DEFws.cell(row = cell.row, column = 11).value
    
    
    diff_of_pyds = opp_passyds - def_pass_yds
    if (diff_of_pyds > 75):
        add_on += 1
    if (diff_of_pyds < -75):
        add_on -= 1
    
    diff_of_ptds = opp_passtds - def_ptds # if the difference of passing touchdowns (compared to league average) is in a certain range, add a certain amt of points
    if (diff_of_ptds >= 1) and (diff_of_ptds<2):
        add_on +=.1
    if (diff_of_ptds <= -1) and (diff_of_ptds>-2):
        add_on -=.1
    if (diff_of_ptds > 2) and (diff_of_ptds<3):
        add_on +=.4
    if (diff_of_ptds <-2) and (diff_of_ptds>-3):
        add_on -=.4
    if (diff_of_ptds > 3):
        add_on += .7
    if (diff_of_ptds < -3):
        add_on -= .7
    
    diff_of_ryds = opp_rushyds - def_rush_yds
    if (diff_of_ryds >=-75) and (diff_of_ryds <=75): 
        add_on = (diff_of_ryds / 25)
    if (diff_of_ryds > 75):
        add_on += 3
    if (diff_of_ryds < -75):
        add_on -= 3
    
    diff_of_rtds = opp_rushtds - def_rush_tds
    if (diff_of_rtds > 5):
        add_on += 1
    if (diff_of_rtds < -5):
        add_on -= 1
    
    final_proj = add_on + base_proj 
    
    
print("The final projection is: ", round(final_proj, 2))

print("Done.")


    